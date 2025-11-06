import os
from typing import Dict, Any
from openpyxl import Workbook, load_workbook


COLUMNS = [
    ("ProcessedAt", 20),
    ("Filename", 30),
    ("Vendor", 30),
    ("Date", 15),
    ("Total", 12),
    ("RawText", 80),
]


class ExcelWriter:
    def __init__(self, path: str) -> None:
        self.path = path
        self._ensure_workbook()

    def _ensure_workbook(self) -> None:
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        if not os.path.exists(self.path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Bills"
            ws.append([name for name, _ in COLUMNS])
            for idx, (_, width) in enumerate(COLUMNS, start=1):
                ws.column_dimensions[chr(64 + idx)].width = width
            wb.save(self.path)

    def append_row(self, info: Dict[str, Any]) -> None:
        wb = load_workbook(self.path)
        ws = wb.active
        ws.append([
            info.get("processed_at", ""),
            info.get("filename", ""),
            info.get("vendor", ""),
            info.get("date", ""),
            info.get("total", ""),
            info.get("raw_text", ""),
        ])
        wb.save(self.path)

    def save(self) -> None:
        # No-op kept for symmetry with potential future batching
        pass


